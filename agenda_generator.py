from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

def create_styled_workshop_agenda(output_file):
    # 1. 建立活頁簿與工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "工作坊議程"

    # 2. 定義表格資料 (依照圖片內容建立)
    data = [
        ["Time", "Content", "Speaker"],
        ["09:00-09:30", "報到", ""],
        ["09:30-09:40", "開場致詞", "王教授"],
        ["09:40-10:05", "邁向 6G 的 AI-RAN 及 O-RAN 趨勢介紹", "劉教授"],
        ["10:05-10:30", "下世代 B5G/6G 專網應用與未來趨勢", "陳教授"],
        ["10:30-10:50", "Break", ""],
        ["10:50-11:20", "從 O-RAN 到 AI-RAN\n智慧通訊的節能應用", "教學團隊"],
        ["11:20-12:00", "O-RAN 環境和各模組化功能介紹", ""],
        ["12:00-13:30", "Lunch", ""],
        ["13:30-14:00", "O-RAN 的市場應用案例", "教學團隊"],
        ["14:00-14:30", "O-RAN OSC 環境建置教學", ""],
        ["14:30-14:50", "Break", ""],
        ["14:50-15:50", "O-RAN OSC 第三方應用程式\nxApps 建置教學", "教學團隊"],
        ["15:50-16:30", "現場討論時間", ""]
    ]

    # 將資料逐列寫入
    for row in data:
        ws.append(row)

    # 3. 合併儲存格 (注意：openpyxl 的列數從 1 開始)
    ws.merge_cells('B2:C2')   # 報到 (跨欄)
    ws.merge_cells('B6:C6')   # Break (跨欄)
    ws.merge_cells('C7:C8')   # 教學團隊 1 (跨列合併)
    ws.merge_cells('B9:C9')   # Lunch (跨欄)
    ws.merge_cells('C10:C11') # 教學團隊 2 (跨列合併)
    ws.merge_cells('B12:C12') # Break (跨欄)
    ws.merge_cells('C13:C14') # 教學團隊 3 (跨列合併)

    # 4. 設定格式：邊框與置中對齊
    thin_border = Side(border_style="thin", color="000000")
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    # wrap_text=True 確保含有 \n 的文字能夠正確換行
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 套用格式到所有儲存格
    for row in ws.iter_rows(min_row=1, max_row=len(data), min_col=1, max_col=3):
        for cell in row:
            cell.border = border
            cell.alignment = alignment

    # 5. 調整欄寬與列高 (美化輸出)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15

    # 針對有換行的列增加高度
    ws.row_dimensions[7].height = 35  # 從 O-RAN 到 AI-RAN...
    ws.row_dimensions[13].height = 35 # O-RAN OSC 第三方應用程式...

    # 6. 儲存檔案
    try:
        wb.save(output_file)
        print(f" 轉換成功！檔案已依圖片格式美化：{output_file}")
    except Exception as e:
        print(f" 發生錯誤：{e}")

# 執行轉換
create_styled_workshop_agenda('agenda.xlsx')
